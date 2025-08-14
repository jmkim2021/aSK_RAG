import os
import re
import math
from typing import Optional, Dict, List

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


def _resolve_data_dir() -> str:
    """데이터 디렉토리 경로를 해석한다.
    우선순위: backend/data(현재 프로젝트) → 사용자가 제시한 절대경로(2가지 표기 모두 시도).
    """
    here = os.path.dirname(__file__)
    local = os.path.join(here, "data")
    if os.path.isdir(local):
        return local

    candidates = [
        "./backend/data",
        "./backend/data",
    ]
    for p in candidates:
        if os.path.isdir(p):
            return p
    return local  # 최후 폴백


# ---- Keyword-inclusion based matching ----


def _normalize_cell(value: Optional[str]) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    return value.strip()


def _tokenize(text: str) -> List[str]:
    """문자열을 토큰 리스트로 변환. 숫자(4.2, 10.3.1), 영문, 한글 단어를 모두 분리.
    소문자 변환 후 반환.
    """
    if not text:
        return []
    text = text.lower()
    tokens = re.findall(r"\d+(?:\.\d+)*|[a-zA-Z]+|[가-힣]+", text)
    return tokens


def _is_zero_cell(text: str) -> bool:
    """셀 값이 '0'으로 간주되는지 판정.
    허용 형태: '0', '0.0', '(0)', '0 %', ' 0 ' 등 괄호/공백/퍼센트 허용.
    숫자 외 문자가 섞여도 전체가 0을 의미하면 True.
    """
    if text is None:
        return False
    s = str(text).strip().lower()
    if not s:
        return False
    # 괄호/공백 제거, 퍼센트/쉼표 제거
    s = s.replace(",", "")
    s = re.sub(r"^[\(\[\{\s]+|[\)\]\}\s]+$", "", s)  # 바깥 괄호/공백 제거
    s = s.rstrip("%")
    # 전부 0 형태인지 확인
    if re.fullmatch(r"0+(?:\.0+)?", s):
        return True
    # 숫자 추출 후 모두 0이면 허용
    nums = re.findall(r"\d+", s)
    if nums and all(int(n) == 0 for n in nums):
        # 소수점 등 비숫자 제거 후도 0 성질 유지
        try:
            val = float(re.sub(r"[^0-9.-]", "", s) or 0)
            return abs(val) == 0.0
        except Exception:
            return True
    return False


STOPWORDS = {
    "and", "the", "of", "in", "to", "for", "on", "by", "with", "at", "from",
    "및", "과", "와", "또는", "또한"
}


def _extract_keywords(text: str) -> List[str]:
    tokens = _tokenize(text)
    keywords: List[str] = []
    for t in tokens:
        if t == "0":
            continue
        if t in STOPWORDS:
            continue
        # 길이가 2 이상이거나 숫자/점 포함 토큰만 키워드로 사용
        if len(t) >= 2 or any(ch.isdigit() for ch in t) or ("." in t):
            keywords.append(t.lower())
    return keywords


def _compare_range_to_binary(gt_ws, md_ws, out_ws, start_col: int, end_col: int, start_row: int, end_row: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            gt_raw = gt_ws.cell(row=r, column=c).value
            md_raw = md_ws.cell(row=r, column=c).value

            gt_text = _normalize_cell(gt_raw)
            md_text = _normalize_cell(md_raw)

            # 둘 다 0(여러 표기 허용) → 일치
            if _is_zero_cell(gt_text) and _is_zero_cell(md_text):
                out_ws.cell(row=r, column=c).value = 1
                continue

            # GT가 비었으면 0
            if not gt_text:
                out_ws.cell(row=r, column=c).value = 0
                continue

            # GT 키워드가 모델 셀에 하나라도 포함되면 1
            gt_keywords = _extract_keywords(gt_text)
            if not gt_keywords:
                out_ws.cell(row=r, column=c).value = 0
                continue
            md_low = (md_text or "").lower()
            matched = any(kw in md_low for kw in gt_keywords)
            out_ws.cell(row=r, column=c).value = 1 if matched else 0


def _ndcg_at_k(values: list[float], k: int) -> float:
    """values는 좌→우 순서의 랭킹 점수(여기서는 0/1).
    DCG@k = sum( rel_i / log2(i+1) ), i:1..k
    IDCG@k = values를 내림차순 정렬했을 때의 DCG@k
    모두 0이면 0 반환.
    """
    if k <= 0:
        return 0.0
    k = min(k, len(values))
    dcg = 0.0
    for i in range(1, k + 1):
        rel = float(values[i - 1] or 0)
        dcg += rel / math.log2(i + 1)
    # IDCG: 상위 k개로 1을 가능한 한 앞에 배치
    num_ones = int(sum(1 for v in values if float(v or 0) > 0))
    m = min(num_ones, k)
    if m == 0:
        return 0.0
    idcg = sum(1.0 / math.log2(i + 1) for i in range(1, m + 1))
    return dcg / idcg if idcg > 0 else 0.0


def _append_ndcg_column(out_ws, start_col: int, end_col: int, start_row: int, end_row: int, k: Optional[int] = None) -> list[float]:
    """지정 범위의 각 행에 대해 NDCG@k를 계산하고, 바로 오른쪽 열에 기록한다.
    k가 None이면 열 개수로 자동 설정.
    """
    write_col = end_col + 1
    k = k or (end_col - start_col + 1)
    # 헤더 기록
    out_ws.cell(row=1, column=write_col).value = f"NDCG@{k}"
    ndcgs: list[float] = []
    for r in range(start_row, end_row + 1):
        row_vals = []
        for c in range(start_col, end_col + 1):
            v = out_ws.cell(row=r, column=c).value
            try:
                row_vals.append(float(v))
            except Exception:
                row_vals.append(0.0)
        score = _ndcg_at_k(row_vals, k)
        out_ws.cell(row=r, column=write_col).value = round(score, 6)
        ndcgs.append(score)
    return ndcgs


def _append_overall_scores(out_ws, start_col: int, end_col: int, start_row: int, end_row: int, ndcgs: list[float]) -> None:
    """시트 하단에 최종 합계/평균 점수 기록.
    - Model Score: Mean NDCG@k (행 평균)
    - Binary Accuracy: C..G 영역의 1 합계 / 총 셀 수
    """
    write_col = end_col + 1  # NDCG 열
    # mean NDCG
    mean_ndcg = sum(ndcgs) / len(ndcgs) if ndcgs else 0.0
    # binary accuracy
    ones = 0.0
    total = 0
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            v = out_ws.cell(row=r, column=c).value
            try:
                ones += float(v)
            except Exception:
                pass
            total += 1
    acc = ones / total if total else 0.0

    # 기록 위치: 데이터 마지막 행 아래 2줄부터
    base_row = end_row + 2
    out_ws.cell(row=base_row, column=2).value = "Model Score (Mean NDCG)"
    out_ws.cell(row=base_row, column=write_col).value = round(mean_ndcg, 6)
    out_ws.cell(row=base_row + 1, column=2).value = "Binary Accuracy"
    out_ws.cell(row=base_row + 1, column=write_col).value = round(acc, 6)


def main():
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl이 설치되어 있지 않습니다. `pip install openpyxl` 후 실행하세요.")

    data_dir = _resolve_data_dir()
    gt_path = os.path.join(data_dir, "ground_truth.xlsx")
    md_path = os.path.join(data_dir, "model_truth.xlsx")
    out_path = os.path.join(data_dir, "ndcg_binary_match.xlsx")

    gt_wb = load_workbook(gt_path)
    md_wb = load_workbook(md_path)

    # 첫 시트 기준
    gt_ws = gt_wb.active
    md_ws = md_wb.active

    # 출력은 ground truth를 베이스로 사용 (나머지 셀 그대로 유지)
    out_wb = gt_wb
    out_ws = out_wb.active

    # C2:G31 → 열 3~7, 행 2~31 (30개 질문)
    start_col, end_col, start_row, end_row = 3, 7, 2, 31
    _compare_range_to_binary(gt_ws, md_ws, out_ws, start_col=start_col, end_col=end_col, start_row=start_row, end_row=end_row)
    # NDCG@k 추가 (k는 열 개수로 자동)
    ndcgs = _append_ndcg_column(out_ws, start_col=start_col, end_col=end_col, start_row=start_row, end_row=end_row, k=None)
    # 최종 합계/평균 점수 추가
    _append_overall_scores(out_ws, start_col=start_col, end_col=end_col, start_row=start_row, end_row=end_row, ndcgs=ndcgs)

    out_wb.save(out_path)
    print(f"완료: {out_path}")


if __name__ == "__main__":
    main()


